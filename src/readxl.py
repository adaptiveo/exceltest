from dataclasses import dataclass
from typing import Any

import openpyxl as opxl

FILENAME = "../excel/test.xlsx"
CUR_NAME_ROW = 3
VAL_NAME_ROW = 5

COME = 0
OUT = 1
CUR_DICT = ["USD", "EUR", "RUR"]
NUM_TYPE = "n"
STR_TYPE = "s"

@dataclass
class CCell:
    type:str
    val:Any

# select workbook
wb = opxl.load_workbook(FILENAME, data_only=True)

# select worksheet
ws = wb.active

max_row_num = ws.max_row
min_row_num = ws.min_row

sum_array = [[0] * 2 for i in range(3)]

shift = COME

for i in range(min_row_num,max_row_num+1):
    curr_cur_cell = ws.cell(i,CUR_NAME_ROW).value
    if curr_cur_cell == None:
        shift = OUT
        continue
    cur_index = CUR_DICT.index(curr_cur_cell)
    curr_num_cell = CCell (ws.cell(i, VAL_NAME_ROW).data_type, ws.cell(i, VAL_NAME_ROW).value)
    if curr_num_cell.type == NUM_TYPE:
        sum_array[cur_index][shift] = sum_array[cur_index][shift] + curr_num_cell.val


print(sum_array)
ws.cell(7,7).value=sum_array[1][1]
wb.save(FILENAME)